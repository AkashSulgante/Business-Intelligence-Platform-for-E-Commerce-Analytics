#!/usr/bin/env python3
"""
Weekly Report Generator for E-commerce BI Platform
Automatically generates comprehensive weekly executive reports in PDF, Excel, and CSV formats.
"""

import os
import sys
import logging
import sqlite3
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
import matplotlib
matplotlib.use("Agg")  # Generate chart files without requiring a desktop GUI.
import matplotlib.pyplot as plt
import seaborn as sns
from typing import Dict, Any, Tuple, List, Optional
import yaml
import json

# For PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# For Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# Initialize logger
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(Path(__file__).parent.parent / "logs" / f"report_generator_{datetime.now().strftime('%Y%m%d')}.log"),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger("weekly_report_generator")

class DataExtractor:
    """Handles extraction of data from the SQLite data warehouse."""

    def __init__(self, db_path: Path):
        self.db_path = db_path
        self.conn = None

    def connect(self):
        """Establish database connection."""
        if self.conn is None:
            self.conn = sqlite3.connect(self.db_path)
            self.conn.row_factory = sqlite3.Row  # Enable column access by name
        return self.conn

    def disconnect(self):
        """Close database connection."""
        if self.conn:
            self.conn.close()
            self.conn = None

    def execute_query(self, query: str, params: Tuple = ()) -> pd.DataFrame:
        """Execute a SQL query and return results as DataFrame."""
        try:
            conn = self.connect()
            df = pd.read_sql_query(query, conn, params=params)
            return df
        except Exception as e:
            logger.error(f"Error executing query: {str(e)}")
            raise

    def get_weekly_data(self, start_date: str, end_date: str) -> Dict[str, pd.DataFrame]:
        """
        Extract data for the specified week.

        Args:
            start_date: Start date in YYYY-MM-DD format
            end_date: End date in YYYY-MM-DD format

        Returns:
            Dictionary of DataFrames for different aspects of the business
        """
        logger.info(f"Extracting data for week: {start_date} to {end_date}")

        data = {}

        # Main sales fact table with dimensions for the week
        sales_query = """
        SELECT
            s.sales_id,
            s.date_id,
            s.customer_id,
            s.product_id,
            s.order_id,
            s.payment_id,
            s.quantity,
            s.unit_price,
            s.total_price,
            s.profit,
            s.freight_value,
            s.return_id,
            s.marketing_id,
            d.date,
            d.day_of_week,
            d.month_name,
            d.quarter,
            d.year,
            c.customer_segment,
            p.product_category_name_english,
            o.order_status,
            pt.payment_type,
            r.return_reason,
            m.channel,
            m.campaign_name
        FROM fact_sales s
        JOIN dim_date d ON s.date_id = d.date_id
        LEFT JOIN dim_customers c ON s.customer_id = c.customer_id
        LEFT JOIN dim_products p ON s.product_id = p.product_id
        LEFT JOIN dim_orders o ON s.order_id = o.order_id
        LEFT JOIN dim_payments pt ON s.payment_id = pt.payment_id
        LEFT JOIN dim_returns r ON s.return_id = r.return_id
        LEFT JOIN dim_marketing m ON s.marketing_id = m.marketing_id
        WHERE d.date BETWEEN ? AND ?
        """

        data['sales'] = self.execute_query(sales_query, (start_date, end_date))
        logger.info(f"Extracted {len(data['sales'])} sales records")

        # Customer metrics
        customer_query = """
        SELECT
            c.customer_id,
            c.customer_segment,
            COUNT(DISTINCT s.order_id) as order_count,
            SUM(s.total_price) as total_spent,
            AVG(s.total_price) as avg_order_value,
            MAX(d.date) as last_purchase_date
        FROM fact_sales s
        JOIN dim_date d ON s.date_id = d.date_id
        JOIN dim_customers c ON s.customer_id = c.customer_id
        WHERE d.date BETWEEN ? AND ?
        GROUP BY c.customer_id, c.customer_segment
        """

        data['customers'] = self.execute_query(customer_query, (start_date, end_date))
        logger.info(f"Extracted metrics for {len(data['customers'])} customers")

        # Product performance
        product_query = """
        SELECT
            p.product_id,
            p.product_category_name_english,
            SUM(s.quantity) as units_sold,
            SUM(s.total_price) as revenue,
            SUM(s.profit) as profit,
            COUNT(DISTINCT s.order_id) as order_count
        FROM fact_sales s
        JOIN dim_date d ON s.date_id = d.date_id
        JOIN dim_products p ON s.product_id = p.product_id
        WHERE d.date BETWEEN ? AND ?
        GROUP BY p.product_id, p.product_category_name_english
        ORDER BY revenue DESC
        """

        data['products'] = self.execute_query(product_query, (start_date, end_date))
        logger.info(f"Extracted performance for {len(data['products'])} products")

        # Daily trends
        daily_query = """
        SELECT
            d.date,
            SUM(s.total_price) as daily_revenue,
            SUM(s.profit) as daily_profit,
            COUNT(DISTINCT s.order_id) as daily_orders,
            COUNT(DISTINCT s.customer_id) as daily_customers
        FROM fact_sales s
        JOIN dim_date d ON s.date_id = d.date_id
        WHERE d.date BETWEEN ? AND ?
        GROUP BY d.date
        ORDER BY d.date
        """

        data['daily_trends'] = self.execute_query(daily_query, (start_date, end_date))
        logger.info(f"Extracted {len(data['daily_trends'])} days of trends")

        # Returns analysis
        returns_query = """
        SELECT
            r.return_reason,
            COUNT(*) as return_count,
            SUM(s.total_price) as returned_value,
            AVG(s.total_price) as avg_return_value
        FROM fact_sales s
        JOIN dim_date d ON s.date_id = d.date_id
        JOIN dim_returns r ON s.return_id = r.return_id
        WHERE d.date BETWEEN ? AND ? AND s.return_id IS NOT NULL
        GROUP BY r.return_reason
        ORDER BY return_count DESC
        """

        data['returns'] = self.execute_query(returns_query, (start_date, end_date))
        logger.info(f"Extracted {len(data['returns'])} return reason categories")

        # Marketing performance
        marketing_query = """
        SELECT
            m.channel,
            m.campaign_name,
            COUNT(DISTINCT s.order_id) as attributed_orders,
            SUM(s.total_price) as attributed_revenue,
            SUM(s.profit) as attributed_profit
        FROM fact_sales s
        JOIN dim_date d ON s.date_id = d.date_id
        JOIN dim_marketing m ON s.marketing_id = m.marketing_id
        WHERE d.date BETWEEN ? AND ? AND s.marketing_id IS NOT NULL
        GROUP BY m.channel, m.campaign_name
        ORDER BY attributed_revenue DESC
        """

        data['marketing'] = self.execute_query(marketing_query, (start_date, end_date))
        logger.info(f"Extracted {len(data['marketing'])} marketing records")

        return data

class MetricsCalculator:
    """Calculates KPIs and generates insights from extracted data."""

    @staticmethod
    def calculate_kpis(data: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
        """
        Calculate key performance indicators from the extracted data.

        Args:
            data: Dictionary of DataFrames from DataExtractor

        Returns:
            Dictionary of calculated KPIs
        """
        kpis = {}

        if 'sales' in data and len(data['sales']) > 0:
            sales_df = data['sales']

            # Revenue KPIs
            kpis['total_revenue'] = sales_df['total_price'].sum()
            kpis['total_profit'] = sales_df['profit'].sum()
            kpis['total_orders'] = sales_df['order_id'].nunique()
            kpis['total_customers'] = sales_df['customer_id'].nunique()
            kpis['total_quantity'] = sales_df['quantity'].sum()

            # Averages
            kpis['avg_order_value'] = kpis['total_revenue'] / kpis['total_orders'] if kpis['total_orders'] > 0 else 0
            kpis['avg_profit_per_order'] = kpis['total_profit'] / kpis['total_orders'] if kpis['total_orders'] > 0 else 0
            kpis['avg_units_per_order'] = kpis['total_quantity'] / kpis['total_orders'] if kpis['total_orders'] > 0 else 0

            # Profit margin
            kpis['profit_margin'] = kpis['total_profit'] / kpis['total_revenue'] if kpis['total_revenue'] > 0 else 0

            # Daily averages
            if 'daily_trends' in data and len(data['daily_trends']) > 0:
                daily_df = data['daily_trends']
                kpis['avg_daily_revenue'] = daily_df['daily_revenue'].mean()
                kpis['avg_daily_orders'] = daily_df['daily_orders'].mean()
                kpis['avg_daily_customers'] = daily_df['daily_customers'].mean()

            # Customer segmentation
            if 'customers' in data and len(data['customers']) > 0:
                cust_df = data['customers']
                if 'customer_segment' in cust_df.columns:
                    segment_dist = cust_df['customer_segment'].value_counts(normalize=True)
                    kpis['customer_segment_distribution'] = segment_dist.to_dict()

                    # Repeat purchase rate (customers with more than one order)
                    repeat_customers = cust_df[cust_df['order_count'] > 1]
                    kpis['repeat_purchase_rate'] = len(repeat_customers) / len(cust_df) if len(cust_df) > 0 else 0

            # Product performance
            if 'products' in data and len(data['products']) > 0:
                prod_df = data['products']
                kpis['top_product_by_revenue'] = prod_df.iloc[0]['product_category_name_english'] if len(prod_df) > 0 else None
                kpis['top_product_revenue'] = prod_df.iloc[0]['revenue'] if len(prod_df) > 0 else 0

                # Product return rate (if returns data available)
                if 'returns' in data and len(data['returns']) > 0:
                    total_sales = sales_df['sales_id'].nunique()
                    returned_sales = sales_df['return_id'].notna().sum()
                    kpis['return_rate'] = returned_sales / total_sales if total_sales > 0 else 0
                else:
                    kpis['return_rate'] = 0

            # Marketing ROI
            if 'marketing' in data and len(data['marketing']) > 0:
                mkt_df = data['marketing']
                # Simplified ROI calculation (would need marketing spend data for accuracy)
                attributed_revenue = mkt_df['attributed_revenue'].sum()
                kpis['marketing_attributed_revenue'] = attributed_revenue
                kpis['marketing_revenue_percentage'] = attributed_revenue / kpis['total_revenue'] if kpis['total_revenue'] > 0 else 0

        return kpis

    @staticmethod
    def generate_insights(data: Dict[str, pd.DataFrame], kpis: Dict[str, Any]) -> List[str]:
        """
        Generate business insights and recommendations based on data and KPIs.

        Args:
            data: Dictionary of DataFrames from DataExtractor
            kpis: Dictionary of calculated KPIs

        Returns:
            List of insight/recommendation strings
        """
        insights = []

        # Revenue insights
        if 'total_revenue' in kpis and 'avg_daily_revenue' in kpis:
            insights.append(f"Weekly revenue totaled ${kpis['total_revenue']:,.2f} with an average daily revenue of ${kpis['avg_daily_revenue']:,.2f}.")

        # Profitability insights
        if 'profit_margin' in kpis:
            margin_pct = kpis['profit_margin'] * 100
            if margin_pct >= 20:
                insights.append(f"Profit margin is strong at {margin_pct:.1f}%, indicating healthy profitability.")
            elif margin_pct >= 10:
                insights.append(f"Profit margin is moderate at {margin_pct:.1f}%, there's room for improvement.")
            else:
                insights.append(f"Profit margin is low at {margin_pct:.1f}%, requiring attention to cost structures.")

        # Customer insights
        if 'repeat_purchase_rate' in kpis:
            repeat_pct = kpis['repeat_purchase_rate'] * 100
            insights.append(f"Repeat purchase rate is {repeat_pct:.1f}%. "
                          f"{'Strong customer loyalty.' if repeat_pct >= 30 else 'Moderate customer retention.' if repeat_pct >= 15 else 'Low repeat purchase rate - consider retention strategies.'}")

        if 'customer_segment_distribution' in kpis:
            segments = kpis['customer_segment_distribution']
            if segments:
                top_segment = max(segments, key=segments.get)
                insights.append(f"The largest customer segment is '{top_segment}' representing {segments[top_segment]*100:.1f}% of customers.")

        # Product insights
        if 'top_product_by_revenue' in kpis and kpis['top_product_by_revenue']:
            insights.append(f"The top-performing product category is '{kpis['top_product_by_revenue']}' "
                          f"with revenue of ${kpis.get('top_product_revenue', 0):,.2f}.")

        # Return insights
        if 'return_rate' in kpis:
            return_pct = kpis['return_rate'] * 100
            insights.append(f"Return rate is {return_pct:.1f}%. "
                          f"{'Acceptable return rate.' if return_pct <= 5 else 'Elevated return rate - review product quality or descriptions.' if return_pct <= 15 else 'High return rate - immediate investigation required.'}")

        # Marketing insights
        if 'marketing_attributed_revenue' in kpis and kpis['total_revenue'] > 0:
            mkt_pct = kpis['marketing_revenue_percentage'] * 100
            insights.append(f"Marketing-attributed revenue represents {mkt_pct:.1f}% of total revenue. "
                          f"{'Strong marketing ROI.' if mkt_pct >= 30 else 'Moderate marketing contribution.' if mkt_pct >= 10 else 'Low marketing attribution - review campaign effectiveness.'}")

        # Daily trends insights
        if 'daily_trends' in data and len(data['daily_trends']) > 0:
            daily_df = data['daily_trends']
            if len(daily_df) >= 2:
                first_day = daily_df.iloc[0]
                last_day = daily_df.iloc[-1]
                revenue_change = ((last_day['daily_revenue'] - first_day['daily_revenue']) / first_day['daily_revenue'] * 100) if first_day['daily_revenue'] > 0 else 0
                insights.append(f"Revenue changed by {revenue_change:+.1f}% from the first to last day of the week.")

        # Add generic insights if specific ones weren't generated
        if len(insights) < 3:
            insights.extend([
                "Monitor inventory levels for top-selling products to prevent stockouts.",
                "Consider A/B testing for email campaigns to improve conversion rates.",
                "Review customer feedback for products with high return rates.",
                "Explore bundling opportunities for complementary products.",
                "Analyze peak shopping times to optimize staffing and promotional timing."
            ])

        return insights[:10]  # Limit to top 10 insights

class ChartGenerator:
    """Generates charts and visualizations for the report."""

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        # Set style
        plt.style.use('seaborn-v0_8')
        sns.set_palette("husl")

    def create_revenue_trend_chart(self, daily_df: pd.DataFrame) -> str:
        """Create a line chart showing daily revenue trend."""
        plt.figure(figsize=(10, 6))
        plt.plot(daily_df['date'], daily_df['daily_revenue'], marker='o', linewidth=2, markersize=4)
        plt.title('Daily Revenue Trend', fontsize=16, fontweight='bold')
        plt.xlabel('Date', fontsize=12)
        plt.ylabel('Revenue ($)', fontsize=12)
        plt.grid(True, alpha=0.3)
        plt.xticks(rotation=45)
        plt.tight_layout()

        chart_path = self.output_dir / "revenue_trend.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        return str(chart_path)

    def create_profit_vs_revenue_chart(self, daily_df: pd.DataFrame) -> str:
        """Create a dual-axis chart showing profit and revenue."""
        fig, ax1 = plt.subplots(figsize=(10, 6))

        color = 'tab:blue'
        ax1.set_xlabel('Date', fontsize=12)
        ax1.set_ylabel('Revenue ($)', color=color, fontsize=12)
        ax1.plot(daily_df['date'], daily_df['daily_revenue'], color=color, marker='o', linewidth=2, label='Revenue')
        ax1.tick_params(axis='y', labelcolor=color)

        ax2 = ax1.twinx()
        color = 'tab:red'
        ax2.set_ylabel('Profit ($)', color=color, fontsize=12)
        ax2.plot(daily_df['date'], daily_df['daily_profit'], color=color, marker='s', linewidth=2, label='Profit')
        ax2.tick_params(axis='y', labelcolor=color)

        plt.title('Daily Revenue vs Profit', fontsize=16, fontweight='bold')
        plt.tight_layout()

        chart_path = self.output_dir / "profit_vs_revenue.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        return str(chart_path)

    def create_sales_by_category_chart(self, products_df: pd.DataFrame) -> str:
        """Create a bar chart showing sales by product category."""
        # Aggregate by category if we have product details
        if 'product_category_name_english' in products_df.columns:
            category_sales = (
                products_df.dropna(subset=['product_category_name_english'])
                .groupby('product_category_name_english')['revenue'].sum()
                .sort_values(ascending=False)
            )
        else:
            category_sales = pd.Series(dtype=float)

        # The bundled source has no category labels; show leading products
        # rather than attempting to plot an empty category series.
        if category_sales.empty:
            category_sales = products_df.set_index('product_id')['revenue'].head(10)

        plt.figure(figsize=(12, 6))
        category_sales.plot(kind='bar')
        plt.title('Sales by Product Category', fontsize=16, fontweight='bold')
        plt.xlabel('Product Category', fontsize=12)
        plt.ylabel('Revenue ($)', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()

        chart_path = self.output_dir / "sales_by_category.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.close()
        return str(chart_path)

    def create_customer_segments_chart(self, customers_df: pd.DataFrame) -> str:
        """Create a pie chart showing customer segment distribution."""
        if 'customer_segment' in customers_df.columns and len(customers_df) > 0:
            segment_counts = customers_df['customer_segment'].dropna().value_counts()
            if segment_counts.empty:
                return None
            plt.figure(figsize=(8, 8))
            plt.pie(segment_counts.values, labels=segment_counts.index, autopct='%1.1f%%', startangle=90)
            plt.title('Customer Segment Distribution', fontsize=16, fontweight='bold')
            plt.tight_layout()

            chart_path = self.output_dir / "customer_segments.png"
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()
            return str(chart_path)
        return None

    def create_returns_pie_chart(self, returns_df: pd.DataFrame) -> str:
        """Create a pie chart showing return reasons."""
        if len(returns_df) > 0 and 'return_reason' in returns_df.columns:
            plt.figure(figsize=(8, 8))
            plt.pie(returns_df['return_count'], labels=returns_df['return_reason'], autopct='%1.1f%%', startangle=90)
            plt.title('Return Reasons Distribution', fontsize=16, fontweight='bold')
            plt.tight_layout()

            chart_path = self.output_dir / "returns_pie.png"
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()
            return str(chart_path)
        return None

    def create_marketing_performance_chart(self, marketing_df: pd.DataFrame) -> str:
        """Create a bar chart showing marketing channel performance."""
        if len(marketing_df) > 0 and 'channel' in marketing_df.columns:
            channel_perf = marketing_df.groupby('channel')['attributed_revenue'].sum().sort_values(ascending=False)
            plt.figure(figsize=(10, 6))
            channel_perf.plot(kind='bar')
            plt.title('Marketing Attributed Revenue by Channel', fontsize=16, fontweight='bold')
            plt.xlabel('Marketing Channel', fontsize=12)
            plt.ylabel('Attributed Revenue ($)', fontsize=12)
            plt.xticks(rotation=45)
            plt.tight_layout()

            chart_path = self.output_dir / "marketing_performance.png"
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()
            return str(chart_path)
        return None

class ReportExporter:
    """Exports the report to various formats."""

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_to_csv(self, data: Dict[str, pd.DataFrame], kpis: Dict[str, Any],
                     insights: List[str], timestamp: str) -> str:
        """Export key metrics and summaries to CSV files."""
        # Create a summary CSV
        summary_data = []
        for kpi, value in kpis.items():
            if isinstance(value, (int, float)):
                summary_data.append({'Metric': kpi, 'Value': value})
            else:
                summary_data.append({'Metric': kpi, 'Value': str(value)})

        summary_df = pd.DataFrame(summary_data)
        csv_path = self.output_dir / f"weekly_report_{timestamp}_executive_summary.csv"
        summary_df.to_csv(csv_path, index=False)
        logger.info(f"Exported executive summary to {csv_path}")

        # Also export detailed data sheets
        for name, df in data.items():
            if df is not None and len(df) > 0:
                df_path = self.output_dir / f"weekly_report_{timestamp}_{name}.csv"
                df.to_csv(df_path, index=False)
                logger.info(f"Exported {name} data to {df_path}")

        return str(csv_path)

    def export_to_excel(self, data: Dict[str, pd.DataFrame], kpis: Dict[str, Any],
                       insights: List[str], chart_paths: Dict[str, str], timestamp: str) -> str:
        """Export the full report to an Excel workbook with multiple sheets."""
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")

        # 1. Executive Summary Sheet
        ws_summary = wb.create_sheet("Executive Summary")
        ws_summary['A1'] = "E-Commerce Weekly Executive Report"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A2'] = f"Period: {timestamp}"
        ws_summary['A2'].font = Font(size=12, italic=True)

        # KPIs table
        row = 4
        ws_summary[f'A{row}'] = "Key Performance Indicators"
        ws_summary[f'A{row}'].font = Font(size=14, bold=True)
        ws_summary.merge_cells(f'A{row}:B{row}')
        row += 1

        # Headers
        ws_summary[f'A{row}'] = "Metric"
        ws_summary[f'B{row}'] = "Value"
        ws_summary[f'A{row}'].font = header_font
        ws_summary[f'B{row}'].font = header_font
        ws_summary[f'A{row}'].fill = header_fill
        ws_summary[f'B{row}'].fill = header_fill
        ws_summary[f'A{row}'].alignment = center_alignment
        ws_summary[f'B{row}'].alignment = center_alignment
        row += 1

        # KPI values
        for kpi, value in kpis.items():
            if isinstance(value, float):
                if 'rate' in kpi or 'margin' in kpi:
                    formatted_value = f"{value:.2%}"
                elif 'revenue' in kpi or 'profit' in kpi or 'value' in kpi:
                    formatted_value = f"${value:,.2f}"
                else:
                    formatted_value = f"{value:,.2f}"
            else:
                formatted_value = str(value)

            ws_summary[f'A{row}'] = kpi.replace('_', ' ').title()
            ws_summary[f'B{row}'] = formatted_value
            row += 1

        # Insights
        row += 2
        ws_summary[f'A{row}'] = "Business Insights & Recommendations"
        ws_summary[f'A{row}'].font = Font(size=14, bold=True)
        ws_summary.merge_cells(f'A{row}:B{row}')
        row += 1

        for insight in insights:
            ws_summary[f'A{row}'] = "•"
            ws_summary[f'B{row}'] = insight
            row += 1

        # 2. Sales Trends Sheet
        if 'daily_trends' in data and len(data['daily_trends']) > 0:
            ws_trends = wb.create_sheet("Daily Trends")
            for r in dataframe_to_rows(data['daily_trends'], index=False, header=True):
                ws_trends.append(r)
            # Apply header style
            for cell in ws_trends[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

        # 3. Product Performance Sheet
        if 'products' in data and len(data['products']) > 0:
            ws_products = wb.create_sheet("Product Performance")
            for r in dataframe_to_rows(data['products'], index=False, header=True):
                ws_products.append(r)
            # Apply header style
            for cell in ws_products[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

        # 4. Customer Analysis Sheet
        if 'customers' in data and len(data['customers']) > 0:
            ws_customers = wb.create_sheet("Customer Analysis")
            for r in dataframe_to_rows(data['customers'], index=False, header=True):
                ws_customers.append(r)
            # Apply header style
            for cell in ws_customers[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

        # 5. Returns Analysis Sheet
        if 'returns' in data and len(data['returns']) > 0:
            ws_returns = wb.create_sheet("Returns Analysis")
            for r in dataframe_to_rows(data['returns'], index=False, header=True):
                ws_returns.append(r)
            # Apply header style
            for cell in ws_returns[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

        # 6. Marketing Performance Sheet
        if 'marketing' in data and len(data['marketing']) > 0:
            ws_marketing = wb.create_sheet("Marketing Performance")
            for r in dataframe_to_rows(data['marketing'], index=False, header=True):
                ws_marketing.append(r)
            # Apply header style
            for cell in ws_marketing[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

        # Save workbook
        excel_path = self.output_dir / f"weekly_report_{timestamp}.xlsx"
        wb.save(excel_path)
        logger.info(f"Exported Excel report to {excel_path}")
        return str(excel_path)

    def export_to_pdf(self, data: Dict[str, pd.DataFrame], kpis: Dict[str, Any],
                     insights: List[str], chart_paths: Dict[str, str], timestamp: str) -> str:
        """Export the report to a PDF document."""
        pdf_path = self.output_dir / f"weekly_report_{timestamp}.pdf"
        doc = SimpleDocTemplate(str(pdf_path), pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER
        )
        story.append(Paragraph("E-Commerce Weekly Executive Report", title_style))
        story.append(Paragraph(f"Period: {timestamp}", styles['Normal']))
        story.append(Spacer(1, 20))

        # KPIs Section
        story.append(Paragraph("Key Performance Indicators", styles['Heading2']))
        story.append(Spacer(1, 12))

        # Create KPIs table
        kpi_data = [["Metric", "Value"]]
        for kpi, value in kpis.items():
            if isinstance(value, float):
                if 'rate' in kpi or 'margin' in kpi:
                    formatted_value = f"{value:.2%}"
                elif 'revenue' in kpi or 'profit' in kpi or 'value' in kpi:
                    formatted_value = f"${value:,.2f}"
                else:
                    formatted_value = f"{value:,.2f}"
            else:
                formatted_value = str(value)
            kpi_data.append([kpi.replace('_', ' ').title(), formatted_value])

        kpi_table = Table(kpi_data, colWidths=[3*inch, 2*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 20))

        # Insights Section
        story.append(Paragraph("Business Insights & Recommendations", styles['Heading2']))
        story.append(Spacer(1, 12))

        for insight in insights:
            story.append(Paragraph(f"• {insight}", styles['Normal']))
            story.append(Spacer(1, 6))

        story.append(PageBreak())

        # Charts Section
        story.append(Paragraph("Visual Analytics", styles['Heading2']))
        story.append(Spacer(1, 12))

        # Add charts if they exist
        chart_order = [
            ('revenue_trend', 'Daily Revenue Trend'),
            ('profit_vs_revenue', 'Daily Revenue vs Profit'),
            ('sales_by_category', 'Sales by Product Category'),
            ('customer_segments', 'Customer Segment Distribution'),
            ('returns_pie', 'Return Reasons Distribution'),
            ('marketing_performance', 'Marketing Performance by Channel')
        ]

        for chart_key, chart_title in chart_order:
            if chart_key in chart_paths and chart_paths[chart_key] and os.path.exists(chart_paths[chart_key]):
                try:
                    ImageReader(chart_paths[chart_key])
                except Exception:
                    logger.warning(f"Skipping unreadable chart file: {chart_paths[chart_key]}")
                    continue
                story.append(Paragraph(chart_title, styles['Heading3']))
                story.append(Spacer(1, 6))
                img = Image(chart_paths[chart_key], width=5*inch, height=3*inch)
                story.append(img)
                story.append(Spacer(1, 20))

        # Build PDF
        doc.build(story)
        logger.info(f"Exported PDF report to {pdf_path}")
        return str(pdf_path)

def get_date_range(days_back: int = 7, reference_date: Optional[datetime] = None) -> Tuple[str, str]:
    """
    Calculate the date range for the report.

    Args:
        days_back: Number of days to look back from yesterday

    Returns:
        Tuple of (start_date, end_date) in YYYY-MM-DD format
    """
    end_date = reference_date or (datetime.now() - timedelta(days=1))
    start_date = end_date - timedelta(days=days_back-1)  # Start date (inclusive)

    return start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')

def main():
    """Main function to generate the weekly report."""
    try:
        # Setup paths
        project_root = Path(__file__).resolve().parent.parent
        db_path = project_root / "data" / "warehouse" / "ecommerce_bi.db"
        reports_dir = project_root / "reports" / "output"
        charts_dir = project_root / "reports" / "output" / "charts"
        logs_dir = project_root / "logs"

        # Create directories if they don't exist
        reports_dir.mkdir(parents=True, exist_ok=True)
        charts_dir.mkdir(parents=True, exist_ok=True)
        logs_dir.mkdir(parents=True, exist_ok=True)

        # Initialize components
        extractor = DataExtractor(db_path)
        calculator = MetricsCalculator()
        chart_gen = ChartGenerator(charts_dir)
        exporter = ReportExporter(reports_dir)

        # Base the report period on the newest warehouse data.  This keeps a
        # historical/demo warehouse useful even when the system date is later.
        latest_date_df = extractor.execute_query("SELECT MAX(date) AS max_date FROM dim_date")
        latest_date = latest_date_df.iloc[0]['max_date'] if not latest_date_df.empty else None
        reference_date = pd.to_datetime(latest_date).to_pydatetime() if pd.notna(latest_date) else None
        start_date, end_date = get_date_range(reference_date=reference_date)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        logger.info(f"Generating report for period: {start_date} to {end_date}")

        # Extract data
        logger.info("Extracting data from data warehouse...")
        data = extractor.get_weekly_data(start_date, end_date)
        extractor.disconnect()

        # Calculate KPIs
        logger.info("Calculating KPIs...")
        kpis = calculator.calculate_kpis(data)

        # Generate insights
        logger.info("Generating business insights...")
        insights = calculator.generate_insights(data, kpis)

        # Generate charts
        logger.info("Generating charts...")
        chart_paths = {}

        if 'daily_trends' in data and len(data['daily_trends']) > 0:
            daily_df = data['daily_trends']
            chart_paths['revenue_trend'] = chart_gen.create_revenue_trend_chart(daily_df)
            chart_paths['profit_vs_revenue'] = chart_gen.create_profit_vs_revenue_chart(daily_df)

        if 'products' in data and len(data['products']) > 0:
            chart_paths['sales_by_category'] = chart_gen.create_sales_by_category_chart(data['products'])

        if 'customers' in data and len(data['customers']) > 0:
            chart_paths['customer_segments'] = chart_gen.create_customer_segments_chart(data['customers'])

        if 'returns' in data and len(data['returns']) > 0:
            chart_paths['returns_pie'] = chart_gen.create_returns_pie_chart(data['returns'])

        if 'marketing' in data and len(data['marketing']) > 0:
            chart_paths['marketing_performance'] = chart_gen.create_marketing_performance_chart(data['marketing'])

        # Export reports
        logger.info("Exporting reports...")
        csv_path = exporter.export_to_csv(data, kpis, insights, timestamp)
        excel_path = exporter.export_to_excel(data, kpis, insights, chart_paths, timestamp)
        pdf_path = exporter.export_to_pdf(data, kpis, insights, chart_paths, timestamp)

        # Print summary
        print("\n" + "="*60)
        print("WEEKLY REPORT GENERATION COMPLETE")
        print("="*60)
        print(f"Report Period: {start_date} to {end_date}")
        print(f"Generated Files:")
        print(f"  CSV Summary:  {csv_path}")
        print(f"  Excel Report: {excel_path}")
        print(f"  PDF Report:   {pdf_path}")
        print(f"  Charts:       {charts_dir}")
        print("="*60)

        # Print key KPIs
        print("\nKEY PERFORMANCE INDICATORS:")
        print(f"  Total Revenue:  ${kpis.get('total_revenue', 0):,.2f}")
        print(f"  Total Profit:   ${kpis.get('total_profit', 0):,.2f}")
        print(f"  Total Orders:   {kpis.get('total_orders', 0):,}")
        print(f"  Total Customers:{kpis.get('total_customers', 0):,}")
        if 'profit_margin' in kpis:
            print(f"  Profit Margin:  {kpis['profit_margin']:.2%}")

        print("\nTOP INSIGHTS:")
        for i, insight in enumerate(insights[:3], 1):
            print(f"  {i}. {insight}")

        return True

    except Exception as e:
        logger.error(f"Report generation failed: {str(e)}", exc_info=True)
        print(f"ERROR: Report generation failed. See logs for details.")
        return False

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
